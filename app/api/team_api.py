"""
Team Management API
Full CRUD for roles and team members.
"""

from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from fastapi.responses import StreamingResponse
from sqlalchemy import select, func, update, delete
from sqlalchemy.ext.asyncio import AsyncSession
from app.core.database import get_db
from app.models.user import User
from app.models.team_role import TeamRole
from app.models.task import Task
from app.api.auth import require_auth
from pydantic import BaseModel
from typing import Optional
import json
import io
import csv
import re

router = APIRouter(prefix="/api/v1/team-mgmt", tags=["team-management"], dependencies=[Depends(require_auth)])


# ── Pydantic Schemas ──

class RoleCreate(BaseModel):
    name: str
    description: Optional[str] = None
    color: Optional[str] = "#3b82f6"
    permissions: Optional[list[str]] = None

class RoleUpdate(BaseModel):
    name: Optional[str] = None
    description: Optional[str] = None
    color: Optional[str] = None
    permissions: Optional[list[str]] = None
    is_default: Optional[bool] = None

class MemberCreate(BaseModel):
    first_name: str
    last_name: Optional[str] = None
    telegram_id: Optional[int] = None
    telegram_username: Optional[str] = None
    phone: Optional[str] = None
    email: Optional[str] = None
    department: Optional[str] = None
    title: Optional[str] = None
    notes: Optional[str] = None
    role_id: Optional[int] = None
    is_admin: bool = False

class MemberUpdate(BaseModel):
    first_name: Optional[str] = None
    last_name: Optional[str] = None
    telegram_username: Optional[str] = None
    phone: Optional[str] = None
    email: Optional[str] = None
    department: Optional[str] = None
    title: Optional[str] = None
    notes: Optional[str] = None
    role_id: Optional[int] = None
    is_admin: Optional[bool] = None
    is_active: Optional[bool] = None
    avatar_url: Optional[str] = None

class InviteMember(BaseModel):
    telegram_username: str
    role_id: Optional[int] = None


# ── Helper ──

def _user_to_dict(u: User, role: Optional[TeamRole] = None, task_stats: dict = None):
    return {
        "id": u.id,
        "telegram_id": u.telegram_id,
        "telegram_username": u.telegram_username,
        "first_name": u.first_name,
        "last_name": u.last_name,
        "full_name": f"{u.first_name or ''} {u.last_name or ''}".strip() or "Unknown",
        "is_admin": u.is_admin,
        "is_active": u.is_active,
        "phone": u.phone,
        "email": u.email,
        "department": u.department,
        "title": u.title,
        "notes": u.notes,
        "avatar_url": u.avatar_url,
        "role_id": u.role_id,
        "role": role.to_dict() if role else None,
        "timezone": u.timezone,
        "task_stats": task_stats or {},
        "created_at": u.created_at.isoformat() if u.created_at else None,
    }


async def _get_task_stats(db: AsyncSession, user_name: str):
    """Get task stats for a user by their name."""
    if not user_name:
        return {"total": 0, "todo": 0, "active": 0, "done": 0}
    result = await db.execute(
        select(Task.status, func.count(Task.id))
        .where(Task.assignee_name == user_name)
        .group_by(Task.status)
    )
    stats = {"total": 0, "todo": 0, "active": 0, "review": 0, "done": 0}
    for status, count in result.all():
        key = status.value if hasattr(status, 'value') else str(status)
        if key == "in_progress":
            key = "active"
        stats[key] = count
        stats["total"] += count
    return stats


# ═══════════════════════════════════════════
# ROLES CRUD
# ═══════════════════════════════════════════

@router.get("/roles")
async def list_roles(db: AsyncSession = Depends(get_db)):
    """List all team roles with member counts."""
    result = await db.execute(select(TeamRole).order_by(TeamRole.sort_order, TeamRole.id))
    roles = result.scalars().all()

    roles_data = []
    for r in roles:
        cnt = await db.execute(select(func.count(User.id)).where(User.role_id == r.id))
        rd = r.to_dict()
        rd["member_count"] = cnt.scalar() or 0
        roles_data.append(rd)

    return {"roles": roles_data}


@router.post("/roles")
async def create_role(data: RoleCreate, db: AsyncSession = Depends(get_db)):
    """Create a new role."""
    # Check unique name
    existing = await db.execute(select(TeamRole).where(TeamRole.name == data.name))
    if existing.scalar_one_or_none():
        raise HTTPException(status_code=400, detail="Role name already exists")

    max_order = await db.execute(select(func.max(TeamRole.sort_order)))
    next_order = (max_order.scalar() or 0) + 1

    role = TeamRole(
        name=data.name,
        description=data.description,
        color=data.color or "#3b82f6",
        permissions=json.dumps(data.permissions) if data.permissions else None,
        sort_order=next_order,
    )
    db.add(role)
    await db.commit()
    await db.refresh(role)
    return role.to_dict()


@router.patch("/roles/{role_id}")
async def update_role(role_id: int, data: RoleUpdate, db: AsyncSession = Depends(get_db)):
    """Update a role."""
    result = await db.execute(select(TeamRole).where(TeamRole.id == role_id))
    role = result.scalar_one_or_none()
    if not role:
        raise HTTPException(status_code=404, detail="Role not found")

    updates = data.model_dump(exclude_unset=True)
    if "permissions" in updates and updates["permissions"] is not None:
        updates["permissions"] = json.dumps(updates["permissions"])

    # If setting as default, unset other defaults
    if updates.get("is_default"):
        await db.execute(update(TeamRole).values(is_default=False))

    for field, value in updates.items():
        setattr(role, field, value)

    await db.commit()
    await db.refresh(role)
    return role.to_dict()


@router.delete("/roles/{role_id}")
async def delete_role(role_id: int, db: AsyncSession = Depends(get_db)):
    """Delete a role. Members lose their role."""
    result = await db.execute(select(TeamRole).where(TeamRole.id == role_id))
    role = result.scalar_one_or_none()
    if not role:
        raise HTTPException(status_code=404, detail="Role not found")

    # Unset role for all members with this role
    await db.execute(update(User).where(User.role_id == role_id).values(role_id=None))
    await db.delete(role)
    await db.commit()
    return {"ok": True, "message": f"Role '{role.name}' deleted."}


# ═══════════════════════════════════════════
# MEMBERS CRUD
# ═══════════════════════════════════════════

@router.get("/members")
async def list_members(db: AsyncSession = Depends(get_db)):
    """List all team members with roles and task stats."""
    result = await db.execute(select(User).order_by(User.created_at.asc()))
    users = result.scalars().all()

    # Load all roles
    roles_result = await db.execute(select(TeamRole))
    roles_map = {r.id: r for r in roles_result.scalars().all()}

    members = []
    for u in users:
        role = roles_map.get(u.role_id) if u.role_id else None
        name = u.first_name or u.telegram_username or "Unknown"
        stats = await _get_task_stats(db, name)
        members.append(_user_to_dict(u, role, stats))

    return {"members": members}


@router.post("/members")
async def create_member(data: MemberCreate, db: AsyncSession = Depends(get_db)):
    """Manually add a new team member."""
    # Generate a placeholder telegram_id if not provided
    tg_id = data.telegram_id
    if not tg_id:
        # Use a negative ID as placeholder for manually added members
        max_result = await db.execute(select(func.min(User.telegram_id)))
        min_id = max_result.scalar() or 0
        tg_id = min(min_id, 0) - 1

    # Check if telegram_id already exists
    existing = await db.execute(select(User).where(User.telegram_id == tg_id))
    if existing.scalar_one_or_none():
        raise HTTPException(status_code=400, detail="A member with this Telegram ID already exists")

    user = User(
        telegram_id=tg_id,
        telegram_username=data.telegram_username,
        first_name=data.first_name,
        last_name=data.last_name,
        phone=data.phone,
        email=data.email,
        department=data.department,
        title=data.title,
        notes=data.notes,
        role_id=data.role_id,
        is_admin=data.is_admin,
    )
    db.add(user)
    await db.commit()
    await db.refresh(user)

    role = None
    if user.role_id:
        rr = await db.execute(select(TeamRole).where(TeamRole.id == user.role_id))
        role = rr.scalar_one_or_none()

    return _user_to_dict(user, role)


# ═══════════════════════════════════════════
# BULK IMPORT
# ═══════════════════════════════════════════

@router.get("/members/template")
async def download_member_template(db: AsyncSession = Depends(get_db)):
    """Download Excel template for bulk member import."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()
    ws = wb.active
    ws.title = "Team Members"

    headers = ["first_name *", "last_name", "email", "phone", "department",
               "job_title", "role", "telegram_username", "is_admin", "notes"]

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid")

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # Example row
    example = ["Sokha", "Chea", "sokha@example.com", "+85512345678",
               "Data Bureau", "Developer", "", "@sokha_chea", "No", "Joined from GDT"]
    for col, val in enumerate(example, 1):
        ws.cell(row=2, column=col, value=val)

    # Column widths
    widths = [15, 15, 25, 18, 20, 20, 15, 20, 10, 25]
    for i, w in enumerate(widths):
        ws.column_dimensions[chr(65 + i)].width = w

    ws.freeze_panes = "A2"

    # Second sheet: available roles
    ws2 = wb.create_sheet("Available Roles")
    ws2.cell(row=1, column=1, value="Role Name").font = Font(bold=True)
    ws2.cell(row=1, column=2, value="Description").font = Font(bold=True)

    result = await db.execute(select(TeamRole).order_by(TeamRole.name))
    roles = result.scalars().all()
    for i, r in enumerate(roles, 2):
        ws2.cell(row=i, column=1, value=r.name)
        ws2.cell(row=i, column=2, value=r.description or "")
    ws2.column_dimensions["A"].width = 20
    ws2.column_dimensions["B"].width = 40

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=team_members_template.xlsx"}
    )


def _parse_import_file(data: bytes, ext: str) -> list[dict]:
    """Parse Excel or CSV into a list of row dicts with row numbers."""
    HEADER_MAP = {
        "first_name": "first_name", "first name": "first_name",
        "last_name": "last_name", "last name": "last_name",
        "email": "email", "phone": "phone",
        "department": "department",
        "job_title": "job_title", "job title": "job_title", "title": "job_title",
        "role": "role",
        "telegram_username": "telegram_username", "telegram username": "telegram_username", "telegram": "telegram_username",
        "is_admin": "is_admin", "admin": "is_admin",
        "notes": "notes",
    }

    def normalize_header(h):
        if not h:
            return ""
        return re.sub(r'[*\s]+', ' ', str(h).strip()).strip().lower()

    rows = []

    if ext in ("xlsx", "xls"):
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
        ws = wb.active
        all_rows = list(ws.iter_rows(values_only=True))
        wb.close()
        if not all_rows:
            return []

        # Map header columns
        raw_headers = [normalize_header(h) for h in all_rows[0]]
        col_map = {}
        for idx, h in enumerate(raw_headers):
            if h in HEADER_MAP:
                col_map[idx] = HEADER_MAP[h]

        if "first_name" not in col_map.values():
            raise ValueError("Missing required column: 'first_name'. Expected headers: first_name, last_name, email, phone, department, job_title, role, telegram_username, is_admin, notes")

        for row_num, row in enumerate(all_rows[1:], start=2):
            # Skip empty rows
            if not any(cell for cell in row):
                continue
            row_dict = {"_row": row_num}
            for idx, field in col_map.items():
                val = row[idx] if idx < len(row) else None
                row_dict[field] = str(val).strip() if val is not None else ""
            rows.append(row_dict)
    else:
        # CSV
        text = data.decode("utf-8-sig", errors="replace")
        reader = csv.DictReader(io.StringIO(text))
        if not reader.fieldnames:
            return []

        col_map = {}
        for h in reader.fieldnames:
            nh = normalize_header(h)
            if nh in HEADER_MAP:
                col_map[h] = HEADER_MAP[nh]

        if "first_name" not in col_map.values():
            raise ValueError("Missing required column: 'first_name'. Expected headers: first_name, last_name, email, phone, department, job_title, role, telegram_username, is_admin, notes")

        for row_num, row in enumerate(reader, start=2):
            # Skip empty rows
            if not any(row.values()):
                continue
            row_dict = {"_row": row_num}
            for orig_h, field in col_map.items():
                val = row.get(orig_h, "")
                row_dict[field] = val.strip() if val else ""
            rows.append(row_dict)

    return rows


@router.post("/members/bulk-import")
async def bulk_import_members(
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
):
    """Bulk import team members from Excel or CSV file."""
    # 1. Validate file type
    filename = file.filename or ""
    ext = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""
    if ext not in ("xlsx", "xls", "csv"):
        raise HTTPException(status_code=400, detail="Only .xlsx, .xls, or .csv files are accepted")

    # 2. Read file (max 5MB)
    file_bytes = await file.read()
    if len(file_bytes) > 5 * 1024 * 1024:
        raise HTTPException(status_code=413, detail="File too large. Maximum 5MB.")

    # 3. Parse rows
    try:
        rows = _parse_import_file(file_bytes, ext)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    except Exception:
        raise HTTPException(status_code=400, detail="Failed to parse file. Ensure it's a valid Excel or CSV file.")

    if not rows:
        return {"total_rows": 0, "imported": 0, "skipped": 0, "failed": 0, "errors": [], "imported_members": []}

    if len(rows) > 500:
        raise HTTPException(status_code=400, detail="Maximum 500 rows per import.")

    # 4. Load existing data for duplicate detection
    existing_emails = set()
    existing_usernames = set()
    result = await db.execute(select(User.email, User.telegram_username))
    for email, username in result.all():
        if email:
            existing_emails.add(email.lower())
        if username:
            existing_usernames.add(username.lower())

    # Load roles for name lookup
    result = await db.execute(select(TeamRole))
    all_roles = result.scalars().all()
    roles_by_name = {r.name.lower(): r for r in all_roles}
    role_names_display = ", ".join(r.name for r in all_roles) if all_roles else "none"

    # Get starting placeholder telegram_id
    min_result = await db.execute(select(func.min(User.telegram_id)))
    min_tg_id = min_result.scalar() or 0
    next_tg_id = min(min_tg_id, 0) - 1

    # 5. Validate and collect
    errors = []
    valid_rows = []
    import_emails = set()
    import_usernames = set()
    skipped = 0

    for row in rows:
        row_num = row.get("_row", "?")
        first_name = row.get("first_name", "").strip()

        # Skip header-like rows
        if first_name.lower() in ("first_name", "first name", "name"):
            skipped += 1
            continue

        # Required: first_name
        if not first_name:
            errors.append({"row": row_num, "message": "first_name is required"})
            continue

        row_errors = []

        # Email validation
        email = row.get("email", "").strip()
        if email:
            if not re.match(r"[^@]+@[^@]+\.[^@]+", email):
                row_errors.append(f"invalid email format '{email}'")
            elif email.lower() in existing_emails or email.lower() in import_emails:
                row_errors.append(f"email '{email}' already exists")

        # Telegram username validation
        tg_user = row.get("telegram_username", "").strip().lstrip("@")
        if tg_user:
            if tg_user.lower() in existing_usernames or tg_user.lower() in import_usernames:
                row_errors.append(f"telegram_username '@{tg_user}' already exists")

        # Role lookup
        role_name = row.get("role", "").strip()
        role_id = None
        if role_name:
            matched_role = roles_by_name.get(role_name.lower())
            if not matched_role:
                row_errors.append(f"role '{role_name}' not found (available: {role_names_display})")
            else:
                role_id = matched_role.id

        if row_errors:
            for err in row_errors:
                errors.append({"row": row_num, "message": err})
            continue

        # Parse is_admin
        is_admin_str = row.get("is_admin", "").strip().lower()
        is_admin = is_admin_str in ("yes", "true", "1")

        # Track for in-file duplicate detection
        if email:
            import_emails.add(email.lower())
        if tg_user:
            import_usernames.add(tg_user.lower())

        valid_rows.append({
            "first_name": first_name,
            "last_name": row.get("last_name", "").strip() or None,
            "email": email or None,
            "phone": row.get("phone", "").strip() or None,
            "department": row.get("department", "").strip() or None,
            "title": row.get("job_title", "").strip() or None,
            "telegram_username": tg_user or None,
            "role_id": role_id,
            "is_admin": is_admin,
            "notes": row.get("notes", "").strip() or None,
            "telegram_id": next_tg_id,
        })
        next_tg_id -= 1

    # 6. Create valid users
    imported_members = []
    for data in valid_rows:
        user = User(
            telegram_id=data["telegram_id"],
            telegram_username=data["telegram_username"],
            first_name=data["first_name"],
            last_name=data["last_name"],
            phone=data["phone"],
            email=data["email"],
            department=data["department"],
            title=data["title"],
            notes=data["notes"],
            role_id=data["role_id"],
            is_admin=data["is_admin"],
            is_active=True,
        )
        db.add(user)
        imported_members.append({"first_name": data["first_name"], "last_name": data["last_name"] or ""})

    if valid_rows:
        await db.commit()

    return {
        "total_rows": len(rows),
        "imported": len(valid_rows),
        "skipped": skipped,
        "failed": len(errors),
        "errors": errors,
        "imported_members": imported_members,
    }


@router.get("/members/{member_id}")
async def get_member(member_id: int, db: AsyncSession = Depends(get_db)):
    """Get single member profile."""
    result = await db.execute(select(User).where(User.id == member_id))
    u = result.scalar_one_or_none()
    if not u:
        raise HTTPException(status_code=404, detail="Member not found")

    role = None
    if u.role_id:
        rr = await db.execute(select(TeamRole).where(TeamRole.id == u.role_id))
        role = rr.scalar_one_or_none()

    name = u.first_name or u.telegram_username or "Unknown"
    stats = await _get_task_stats(db, name)
    return _user_to_dict(u, role, stats)


@router.patch("/members/{member_id}")
async def update_member(member_id: int, data: MemberUpdate, db: AsyncSession = Depends(get_db)):
    """Update a team member's profile."""
    result = await db.execute(select(User).where(User.id == member_id))
    user = result.scalar_one_or_none()
    if not user:
        raise HTTPException(status_code=404, detail="Member not found")

    for field, value in data.model_dump(exclude_unset=True).items():
        setattr(user, field, value)

    await db.commit()
    await db.refresh(user)

    role = None
    if user.role_id:
        rr = await db.execute(select(TeamRole).where(TeamRole.id == user.role_id))
        role = rr.scalar_one_or_none()

    return _user_to_dict(user, role)


@router.delete("/members/{member_id}")
async def delete_member(member_id: int, db: AsyncSession = Depends(get_db)):
    """Remove a team member."""
    result = await db.execute(select(User).where(User.id == member_id))
    user = result.scalar_one_or_none()
    if not user:
        raise HTTPException(status_code=404, detail="Member not found")

    name = f"{user.first_name or ''} {user.last_name or ''}".strip()
    await db.delete(user)
    await db.commit()
    return {"ok": True, "message": f"Member '{name}' removed."}


@router.post("/invite")
async def invite_member(data: InviteMember, db: AsyncSession = Depends(get_db)):
    """Invite a member by Telegram username (creates placeholder entry)."""
    username = data.telegram_username.lstrip("@")

    # Check if username already exists
    existing = await db.execute(
        select(User).where(User.telegram_username == username)
    )
    if existing.scalar_one_or_none():
        raise HTTPException(status_code=400, detail=f"@{username} is already a team member")

    # Create placeholder — will be updated when they message the bot
    max_result = await db.execute(select(func.min(User.telegram_id)))
    min_id = max_result.scalar() or 0
    tg_id = min(min_id, 0) - 1

    user = User(
        telegram_id=tg_id,
        telegram_username=username,
        first_name=username,
        role_id=data.role_id,
        is_active=True,
    )
    db.add(user)
    await db.commit()
    await db.refresh(user)

    role = None
    if user.role_id:
        rr = await db.execute(select(TeamRole).where(TeamRole.id == user.role_id))
        role = rr.scalar_one_or_none()

    return _user_to_dict(user, role)
